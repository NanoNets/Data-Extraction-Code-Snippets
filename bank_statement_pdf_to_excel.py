import PyPDF2
import re
import openpyxl
from openpyxl.styles import Font, Alignment

def extract_text_from_pdf(pdf_path):
    with open(pdf_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        text = ''
        for page in reader.pages:
            text += page.extract_text()
    return text

def parse_transactions(text):
    transactions = []
    lines = text.split('\n')
    current_date = None
    
    for line in lines:
        date_match = re.match(r'(\d{2}/\d{2}/\d{4})', line)
        if date_match:
            current_date = date_match.group(1)
            continue
        
        transaction_match = re.match(r'(.+?)\s+([-+]?\$?\d+\.\d{2})\s+([-+]?\$?\d+\.\d{2})$', line)
        if transaction_match and current_date:
            description, amount, balance = transaction_match.groups()
            transactions.append({
                'date': current_date,
                'description': description.strip(),
                'amount': amount,
                'balance': balance
            })
    
    return transactions

def create_excel_from_transactions(transactions, excel_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Statement"

    headers = ['Date', 'Description', 'Amount', 'Balance']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row, transaction in enumerate(transactions, start=2):
        ws.cell(row=row, column=1, value=transaction['date'])
        ws.cell(row=row, column=2, value=transaction['description'])
        ws.cell(row=row, column=3, value=transaction['amount'])
        ws.cell(row=row, column=4, value=transaction['balance'])

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_path)
    print(f"Excel file created: {excel_path}")

def main():
    pdf_path = 'bank_statement.pdf'  # Replace with your PDF file path
    excel_path = 'bank_statement_output.xlsx'  # Replace with desired output Excel file path

    # Extract text from PDF
    pdf_text = extract_text_from_pdf(pdf_path)

    # Parse transactions from extracted text
    transactions = parse_transactions(pdf_text)

    # Create Excel file from parsed transactions
    create_excel_from_transactions(transactions, excel_path)

if __name__ == "__main__":
    main()